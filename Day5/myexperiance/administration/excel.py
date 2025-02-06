
import openpyxl

def read_excel(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data

def write_excel(file_path, sheet_name, data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for row in data:
        sheet.append(row)
    workbook.save(file_path)
